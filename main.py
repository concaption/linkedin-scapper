import traceback
from selenium.webdriver.edge.options import Options
import threading
import os
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium import webdriver
import chromedriver_autoinstaller
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl, datetime
import time
import psutil
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from collections import defaultdict
import pandas as pd
import time,csv
from unidecode import unidecode
from datetime import date
from dateutil.relativedelta import relativedelta
 
class Linkedin_Scraper():
    def __init__(self):
        self.fileName1 = 'Linkedin_Data.csv'
        self.fileName2 = 'Linkedin_Urls.xlsx'
        self.fileName3 = 'Login_Password_LinkedIn.xlsx'
        self.fileName4 = 'Linkedin_Check_Urls.txt'
 
        self.edge_options = webdriver.EdgeOptions()
        self.edge_options.add_argument("--disable-notifications")
        self.edge_options.add_argument("--disable-gpu")
        self.edge_options.add_argument("--disable-popup-blocking")
        self.edge_options.add_argument("--profile-directory=Default")
        self.edge_options.add_argument("--ignore-certificate-errors")
        self.edge_options.add_argument("--disable-plugins-discovery")
        self.edge_options.add_experimental_option("debuggerAddress", "localhost:9515")
 
    def startEdge(self):
        return os.system('"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe" --remote-debugging-port=9515')
 
    def closeBrowser(self):
        try:
            PROCNAME = "msedgedriver.exe"
            userName = os.getlogin()
            for proc in psutil.process_iter():
                if proc.name() == PROCNAME or proc.name() == 'msedge.exe':
                    if str(userName) in str(proc.username()):
                        print(str(proc.name()))
                        print(proc.username())
                        proc.kill()
        except Exception as ex:
            print(str(ex))
 
 
    def startScraping(self):
        location = os.getcwd()
        path = os.path.join(location,'msedgedriver.exe')
        print(path)
        time.sleep(2)
        self.closeBrowser()
        time.sleep(2)
        th = threading.Thread(target=self.startEdge, args=())
        th.daemon = True
        th.start()
        time.sleep(3)
        driver = webdriver.Edge(options=self.edge_options)
        return driver
 
    def read_Urls(self):
        file_path = self.fileName2
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        num_rows = worksheet.max_row
        allRows = []
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        url_column_index = header_row.index('Linkedin Urls')
        for row in range(2, num_rows + 1):
            row_data = worksheet[row]
            if row_data[url_column_index].value is None:
                break
            data = [
                row_data[url_column_index].value,
            ]
            allRows.append(data)
        workbook.close()
        return allRows
 
    def ReadData(self):
        file_path = self.fileName3
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        num_rows = worksheet.max_row
        allRows = []
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        login_column_index = header_row.index('Login')
        Password_column_index = header_row.index('Password')
        for row in range(2, num_rows + 1):
            row_data = worksheet[row]
            if row_data[login_column_index].value is None:
                break
            data = [
                row_data[login_column_index].value,
                row_data[Password_column_index].value,
            ]
            allRows.append(data)
        workbook.close()
        return allRows
    
    def other_extractor(self,driver):
        time.sleep(3)
        try: newCompany = driver.find_element(By.XPATH,'//ul/li[contains(@id,"profilePagedListComponent")]//ul/li[@class="pvs-list__paged-list-item  pvs-list__item--one-column"]/ancestor::li[2]//div[@class="display-flex flex-wrap align-items-center full-height"]//span[1]').text.replace('Â','A')
        except: newCompany = ''
        try: timeTags = driver.find_elements(By.XPATH,'//ul/li[contains(@id,"profilePagedListComponent")]//ul/li[@class="pvs-list__paged-list-item  pvs-list__item--one-column"]//ancestor::li[2]//div[@class="display-flex flex-wrap align-items-center full-height"]/following-sibling::span[1]/span[1]')[1:]
        except: timeTags = ''
        try: worksTags = driver.find_elements(By.XPATH,'//ul/li[contains(@id,"profilePagedListComponent")]//ul/li[@class="pvs-list__paged-list-item  pvs-list__item--one-column"]/ancestor::li[2]//div[@class="display-flex flex-wrap align-items-center full-height"]//span[@aria-hidden="true"]')[1:]
        except: worksTags = ''
        try: LocationTags = driver.find_elements(By.XPATH,'//ul/li[contains(@id,"profilePagedListComponent")]//ul/li[@class="pvs-list__paged-list-item  pvs-list__item--one-column"]/ancestor::li[2]//div[@class="display-flex flex-wrap align-items-center full-height"]/following-sibling::span[2]/span[1]')[-1].text.replace('Â','A')
        except: LocationTags = ''
        allRows = []
        for tim,work in zip(timeTags,worksTags):
            clearWork = work.text
            cleartime = tim.text
            splittedTime = cleartime.split('-')
            firstDateTag = splittedTime[0].strip().replace('Â','A')
            lastDateTag = splittedTime[-1].split('·')[0].strip().replace('Â','A')
            try:
                from datetime import datetime
                input_date = datetime.strptime(firstDateTag, "%b %Y")
                firstDate = input_date.date()
            except: firstDate = firstDateTag
            lastDate = lastDateTag 
            # timeFrame = ''
            try:
                if lastDateTag.isalpha(): pass
                else:
                    try: 
                        from datetime import datetime
                        input_date2 = datetime.strptime(lastDateTag, "%b %Y")
                        lastDate = input_date2.date()
                    except:
                        lastDate = lastDateTag
                timeFrame = f'{firstDate} to {lastDate}'
            except: 
                lastDate = lastDateTag
                timeFrame = f'{firstDate} to {lastDate}'
            if all([clearWork,newCompany,timeFrame,LocationTags]):
                newRow2 = [unidecode(clearWork),newCompany,timeFrame,LocationTags]
                allRows += (newRow2)
        return allRows
        
    def extract_Data(self,driver,url):
        print()
        try:
            with open(file=self.fileName4,mode='r',encoding='utf-8') as f:
                readData = f.readlines()
                allData = [data.strip() for data in readData]
            if url not in allData:
                # print(url)
                with open(file=self.fileName4,mode='a',encoding='utf-8') as f:
                    f.write(url + '\n')
                searchUrl = f'{url}/details/experience/'
                driver.get(searchUrl)
                time.sleep(6)
                try:
                    checkTags = driver.find_elements(By.XPATH,'//li[@class="pvs-list__paged-list-item artdeco-list__item pvs-list__item--line-separated pvs-list__item--one-column"]//div[@class="display-flex flex-wrap align-items-center full-height"]/following-sibling::span')
                    allLocation = ' '.join([tag.get_attribute('textContent') for tag in checkTags])
                    if 'Lithuania' in allLocation:
                        time.sleep(2)
                        try: titleTag = driver.find_element(By.XPATH,'//div[@class="artdeco-entity-lockup__title ember-view"]').text.replace('Â','A')
                        except: titleTag = ''
                        if titleTag:
                            title = unidecode(titleTag) if titleTag else ''
                            time.sleep(2)
                            firsExperinesTag = driver.find_elements(By.XPATH,'//li[@class="pvs-list__paged-list-item artdeco-list__item pvs-list__item--line-separated pvs-list__item--one-column"]')
                            allExperinessTag = []
                            # print(firsExperinesTag)
                            time.sleep(1)
                            for box in firsExperinesTag:
                                try: workTitle = box.find_element(By.XPATH,'.//div[@class="display-flex flex-wrap align-items-center full-height"]//span[1]').text.replace('Â','A')
                                except: workTitle = ''
                                try: company_Name = box.find_element(By.XPATH,'.//div[@class="display-flex flex-wrap align-items-center full-height"]/following-sibling::span[1]/span[1]').text.replace('Â','A')
                                except: company_Name = ''
                                try: timeFrameTag = box.find_element(By.XPATH,'.//div[@class="display-flex flex-wrap align-items-center full-height"]/following-sibling::span[2]/span[1]').text.replace('Â','A')
                                except: timeFrameTag = ''
                                try: location = box.find_element(By.XPATH,'.//div[@class="display-flex flex-wrap align-items-center full-height"]/following-sibling::span[3]/span[1]').text.replace('Â','A')
                                except: location = ''
                                if all([workTitle,company_Name,timeFrameTag,location]):
                                    splittedTime = timeFrameTag.split('-')
                                    firstDateTag = splittedTime[0].strip()
                                    lastDateTag = splittedTime[-1].split('·')[0].strip()
                                    try: 
                                        from datetime import datetime
                                        input_date = datetime.strptime(unidecode(firstDateTag), "%b %Y")
                                        firstDate = input_date.date()
                                    except: firstDate = firstDateTag
                                    lastDate = lastDateTag
                                    # timeFrame = ''
                                    try:
                                        if lastDateTag.isalpha(): pass
                                        else:
                                            try:
                                                from datetime import datetime
                                                input_date2 = datetime.strptime(unidecode(lastDateTag), "%b %Y")
                                                lastDate = input_date2.date()
                                            except: lastDate = lastDateTag
                                        timeFrame = f'{firstDate} to {lastDate}'
                                    except: 
                                        lastDate = lastDateTag
                                        timeFrame = f'{firstDate} to {lastDate}'
                                    row = [workTitle,company_Name,timeFrame,location]
                                    # print(row)
                                    allExperinessTag += (row)
                            try: otherExp = driver.find_elements(By.XPATH,'//ul/li[contains(@id,"profilePagedListComponent")]//ul/li[@class="pvs-list__paged-list-item  pvs-list__item--one-column"]')
                            except: otherExp = ''
                            if otherExp: 
                                newRows2 = self.other_extractor(driver)
                            else: newRows2 = []
                            newRow = [title,url] + allExperinessTag + newRows2
                            # print(newRow)
                            print(f'[INFO] Getting data from this profile:- {title}')
                            self.saveData(newRow)
                        else:pass
                    else:pass
                except: pass
            else: pass
        except: pass
    def saveData(self,row):
        with open(file=self.fileName1,mode='a',newline='',encoding='utf-8') as f:
            csv.writer(f).writerow(row)
    
    def header(self):
        header = ['Profile Name','Profile URL','Work Title-1', 'Company Name-1', 'Time Frame-1', 'Location-1', 'Work Title-2', 'Company Name-2', 'Time Frame-2', 'Location-2', 'Work Title-3', 'Company Name-3', 'Time Frame-3', 'Location-3', 'Work Title-4', 'Company Name-4', 'Time Frame-4', 'Location-4', 'Work Title-5', 'Company Name-5', 'Time Frame-5', 'Location-5', 'Work Title-6', 'Company Name-6', 'Time Frame-6', 'Location-6', 'Work Title-7', 'Company Name-7', 'Time Frame-7', 'Location-7', 'Work Title-8', 'Company Name-8', 'Time Frame-8', 'Location-8', 'Work Title-9', 'Company Name-9', 'Time Frame-9', 'Location-9', 'Work Title-10', 'Company Name-10', 'Time Frame-10', 'Location-10']
        with open(file=self.fileName1,mode='w',newline='',encoding='utf-8') as f:
            csv.writer(f).writerow(header)
    
    def Sign_Process(self,driver2,Email,Password):
        userMail = Email.strip()
        userPassword = Password.strip()
        try:
            time.sleep(2)
            sign_In = driver2.find_element(By.XPATH,'//input[@id="username"]')
            time.sleep(1)
            sign_In.click()
            time.sleep(1)
            sign_In.clear()
            time.sleep(2)
            sign_In.send_keys(userMail)
            time.sleep(3)
            sign_Password = driver2.find_element(By.XPATH,'//input[@id="password"]')
            time.sleep(1)
            sign_Password.click()
            time.sleep(1)
            sign_Password.clear()
            time.sleep(2)
            sign_Password.send_keys(userPassword)
            time.sleep(2)
            submit_Login = driver2.find_element(By.XPATH,'//button[@aria-label="Sign in"]')
            time.sleep(2)
            submit_Login.click()
            time.sleep(7)
        except: pass
        try: checkPin = driver2.find_element(By.XPATH,'//input[@name="pin"]')
        except: checkPin = ''
        if checkPin: time.sleep(40)
        else: pass
 
    def logout(self,driver):
        driver.get('https://www.linkedin.com/m/logout/')
        time.sleep(3)
        print("\nLogging out\n")
    
    def run(self):
        if self.fileName1 not in os.listdir(): self.header()
        if self.fileName2 not in os.listdir(): open(file=self.fileName2,mode='a').close()
        if self.fileName4 not in os.listdir(): open(file=self.fileName4,mode='a').close()
        if self.fileName3 not in os.listdir():
            workbook = openpyxl.Workbook()
            workbook.save(self.fileName3)
        driver = self.startScraping()
        time.sleep(4)
        # driver.get('https://www.linkedin.com/login')
        driver.get('https://www.linkedin.com/checkpoint/lg/sign-in-another-account?')
        time.sleep(4)
        allData = self.ReadData()
        allLogin = []
        for data in allData:
            userEmail = data[0]
            userPassword = data[1]
            makeDict = {'username':userEmail,'password':userPassword}
            allLogin.append(makeDict)
        try: sign_In = driver.find_element(By.XPATH,'//input[@id="username"]')
        except: sign_In = ''
        if sign_In:
            firstLogin = allLogin[0]
            self.Sign_Process(driver,firstLogin["username"], firstLogin["password"])
        else: pass
        profiles = self.read_Urls()
        profiles_processed = 0
        profiles_per_iteration = 60
        if profiles:
            for index,profile in enumerate(profiles,start=1):
                self.extract_Data(driver,profile[0])
                profiles_processed += 1
                if profiles_processed % profiles_per_iteration == 0:
                    self.logout(driver)
                    driver.get('https://www.linkedin.com/checkpoint/lg/sign-in-another-account?')
                    time.sleep(4)
                    current_index = (profiles_processed // profiles_per_iteration - 1) % len(allLogin)
                    current_credentials = allLogin[current_index]
                    self.Sign_Process(driver,current_credentials["username"], current_credentials["password"])
                    driver.get('https://www.linkedin.com')
                    time.sleep(3)
                    try: checkXpath = driver.find_element(By.XPATH,'//span[@title="Home"]')
                    except: checkXpath = ''
                    if checkXpath:
                        pass
                    else:
                        self.Sign_Process(driver,current_credentials["username"], current_credentials["password"])
                    time.sleep(2)
        else:
            print(f'[INFO] Please insert some urls in empty csv, for extract data!')
 
if __name__ == '__main__':
    beenVerified = Linkedin_Scraper()
    beenVerified.run()
 
 
 
 